'''交通費計算功能'''
import tkinter as tk
import openpyxl
import START
import homepage

xl = openpyxl.load_workbook('DB.xlsx', data_only=True)
sh = xl['出席表']
in_name = ''
in_school = ''
in_date = ''
in_num = ''
inbox = ''
viewbox = ''
    
class fare(tk.Frame):
    def __init__(self, master):
        global in_name, in_school, in_date, in_num, inbox, viewbox

        tk.Frame.__init__(self, master)

        # 輸入資料之變數
        in_name = tk.StringVar()
        in_school =tk.StringVar()
        in_date = tk.StringVar()
        in_num = tk.StringVar()
        
        ### 輸入的欄位 ###
        f1 = tk.LabelFrame(self, text = '新增輸入')
        f1.pack()
        a1 = tk.Label(f1, text='姓名')
        a2 = tk.Entry(f1, textvariable = in_name)

        b1 = tk.Label(f1, text='學校')
        b2 = tk.Entry(f1, textvariable = in_school)

        c1 = tk.Label(f1, text='日期')
        c2 = tk.Entry(f1, textvariable = in_date)

        d1 = tk.Label(f1, text='節次')
        d2 = tk.Entry(f1, textvariable = in_num)

        btn = tk.Button(f1, text='輸入', command = self.in_statu)

        a1.grid(column=0, row=0, padx = 10, pady = 10)
        a2.grid(column=1, row=0, padx = 10, pady = 10)
        b1.grid(column=0, row=1, padx = 10, pady = 10)
        b2.grid(column=1, row=1, padx = 10, pady = 10)
        c1.grid(column=0, row=2, padx = 10, pady = 10)
        c2.grid(column=1, row=2, padx = 10, pady = 10)
        d1.grid(column=0, row=3, padx = 10, pady = 10)
        d2.grid(column=1, row=3, padx = 10, pady = 10)
        btn.grid(column=0, row=4, columnspan = 2, padx = 10)


        ### 顯示的欄位 ###
        f2 = tk.LabelFrame(self, text = '資料顯示', padx = 10)
        f2.pack()

        inbox = tk.Listbox(f2, width = 40, height = 15)
        viewbox = tk.Listbox(f2, width = 40, height = 15)
        btn = tk.Button(f2, text='移除', command = self.in_delete)
        btn2 = tk.Button(f2, text='確定新增', command = self.intodata)
        btn3 = tk.Button(f2, text='計算交通費')
        

        inbox.grid(column=0, row=0, columnspan = 2, padx = 10, pady = 10)
        viewbox.grid(column=2, row=0, columnspan = 2, padx = 10, pady = 10)
        btn.grid(column=0, row=1, padx = 10, pady = 10)
        btn2.grid(column=1, row=1, padx = 10, pady = 10)
        btn3.grid(column=2, row=1, columnspan = 2, padx = 10, pady = 10)

        tk.Button(self, text='回主頁'\
            , command = lambda:master.switch_frame(homepage.homepage)).pack()

        self.preview()

    def preview(self):
        global sh, xl, viewbox
        
        xl = openpyxl.load_workbook('DB.xlsx', data_only=True)
        sh = xl['出席表']
        viewbox.delete(0, tk.END)
        for i in sh:
            ss = ''
            for j in i:
                if j.value == None:
                    ss += f'|{' ':^10}'
                else:
                    ss += f'|{j.value:^10}'
            viewbox.insert(tk.END, ss)

    # 輸入按鈕的事件
    def in_statu(self):
        global sh, xl, in_name, in_school, in_date, in_num, inbox, viewbox
        ss = [in_name.get(), in_school.get(), in_date.get(), in_num.get()]
        print(ss)
        #sh.append(s)
        #xl.save('小幫手資料庫.xlsx') ### 寫入
        in_date.set('')
        in_num.set('')
        inbox.insert(tk.END, ss)

    # 刪除按鈕的事件
    def in_delete(self):
        global inbox
        n, = inbox.curselection()
        inbox.delete(n)

    # 確定新增按鈕的事件
    def intodata(self):
        global sh, xl, in_date, in_num, inbox, viewbox
        
        xl = openpyxl.load_workbook('DB.xlsx', data_only=True)
        sh = xl['出席表']

        for s in inbox.get(0, tk.END):
            sh.append(s)
        xl.save('DB.xlsx') ### 寫入
        inbox.delete(0, tk.END)
        self.preview()


